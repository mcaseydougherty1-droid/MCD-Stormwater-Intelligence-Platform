from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_workbook(path: Path, parcels: pd.DataFrame, pipeline: pd.DataFrame, docs: pd.DataFrame, inspections: pd.DataFrame):
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dashboard = pd.DataFrame({
            "Metric": ["Total Parcels", "Target Parcels", "A-Tier Prospects", "Estimated Pipeline Revenue"],
            "Value": [len(parcels), int(parcels["Target_Flag"].sum()), int((pipeline["Pipeline_Tier"].astype(str)=="A").sum()), float(pipeline["Estimated_Annual_Revenue"].sum())],
        })
        dashboard.to_excel(writer, sheet_name="Dashboard", index=False)
        parcels.to_excel(writer, sheet_name="Master Inventory", index=False)
        pipeline.to_excel(writer, sheet_name="Ranked Sales Pipeline", index=False)
        docs.to_excel(writer, sheet_name="Document Crosswalk", index=False)
        inspections.to_excel(writer, sheet_name="Inspection Schedule", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        if ws.max_row > 1 and ws.max_column > 1:
            ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            tab = Table(displayName=ws.title.replace(" ", "_").replace("-", "_")[:25], ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)
    wb.save(path)
